import openpyxl as xl
import numpy as np
import pandas as pd
import pickle

wb = xl.load_workbook(filename='../data/21/BingRenTADboundaries.xlsx', read_only=True, data_only=True)
# data_only does not copy formulae

cells = [k for k in wb.sheetnames if k != 'BL']
boundaries = {k: [] for k in cells}
for cell in cells:
    ws = wb[cell]
    boundaries[cell] = [[k.value for k in row] for row in ws.iter_rows()]

all_tads = []

master = []
for cell, bounds in boundaries.items():
    bounds = sorted(bounds, key=lambda x: (x[0], x[1]))  # sort by chromosome then starting pos
    tads = [[cell] + bounds[i] + bounds[i+1] + ['unvisited'] for i in range(len(bounds) - 1) if bounds[i][0] == bounds[i + 1][0]]
    master += tads

# Unexpanded conservation check
# master_concat = [[k[0]] + ['{0}_{1}_{2}-{0}_{3}_{4}'.format(k[1], k[2], k[3], k[5], k[6])] for k in master]
# master_presence = {}
#
# for line in master_concat:
#     master_presence[line[1]] = {cell: False for cell in cells}
#
# for line in master_concat:
#     master_presence[line[1]][line[0]] = True
#
# master_presence = pd.DataFrame(master_presence).transpose()
#
# with open('../data/21/pickle/pairwise_presence.unexpanded.pickle', 'wb') as f:
#     pickle.dump(master_presence, f, pickle.HIGHEST_PROTOCOL)
#
# exit()

#######################
# ^^ w/o  expansion
# vv with expansion
# uncomment the appropriate section
#######################

# Expanded conservation check


def contains(TAD1, TAD2):
    if TAD1[0][0] != TAD2[0][0]:    # check if on the same chromosome
        return False
    if TAD2[0][1] >= TAD1[0][1] and TAD2[1][2] <= TAD1[1][2]:
        return True
    return False


def expandTAD(TAD, expansion):

    bound1 = [TAD[0][0], TAD[0][1] - expansion, TAD[0][2] - expansion]
    bound2 = [TAD[1][0], TAD[1][1] + expansion, TAD[1][2] + expansion]
    return bound1, bound2


master_presence, idx = [], []
for i, line in enumerate(master):
    if line[-1] == 'visited':
        continue
    presence = {cell: False for cell in cells}
    cell = line[0]
    master[i][-1] = 'visited'
    tad1 = None
    try:
        tad1 = ([int(k) for k in line[1:4]], [int(k) for k in line[4:-1]])    #for chroms 1-22
    except:
        tad1 = ([line[1]] + [int(k) for k in line[2:4]], [line[4]] + [int(k) for k in line[5:-1]]) #for chroms X, Y
    tad1 = expandTAD(tad1, 0)
    presence[cell] = True
    for j, check in enumerate(master):
        try:
            tad2 = ([int(k) for k in check[1:4]], [int(k) for k in check[4:-1]])
        except:
            tad2 = ([check[1]] + [int(k) for k in check[2:4]], [check[4]] + [int(k) for k in check[5:-1]])
        if contains(tad1, tad2):
            presence[check[0]] = True
            master[j][-1] = 'visited'

    pair = '{0}_{1}_{2}-{0}_{3}_{4}'.format(line[1], line[2], line[3], line[5], line[6])
    master_presence.append(presence)
    idx.append(pair)
master_presence = pd.DataFrame(master_presence, index=idx)
with open('../data/21/pickle/pairwise_presence.unexpandedcont.pickle', 'wb') as f:
    pickle.dump(master_presence, f, pickle.HIGHEST_PROTOCOL)

